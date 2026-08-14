# -*- coding: utf-8 -*-
"""
Обратный ход: BPMN 2.0 XML -> таблица.

Зеркало core.build_xml. Чистая логика без ввода-вывода и без сети.
Единственная точка входа: to_table(xml: str | bytes, ...) -> TableResult

Таблица — проекция BPMN с потерями: координаты, расширения движков,
аннотации и чёрные пулы в неё не укладываются. Всё потерянное попадает
в issues, молча ничего не выбрасывается.
"""
from __future__ import annotations

import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Any

from .core import (EVENT_DEFS, NODE_TYPES, NS, SUBPROCESS_TYPES,
                   ConvertError, Issue)

B = f"{{{NS['bpmn']}}}"

# Порядок колонок совпадает с шаблоном template.xlsx
COLUMN_ORDER = [
    "id", "name", "type", "event_def", "pool", "lane", "parent_id",
    "attached_to", "cancel_activity", "marker", "next", "message_to",
    "data_in", "data_out", "documentation",
]
# Эти колонки выводятся всегда, даже пустыми
ALWAYS = {"id", "name", "type", "pool", "next"}

EVENT_DEF_BY_TAG = {v: k for k, v in EVENT_DEFS.items()}

# Дети контейнера, которые не являются узлами и о которых молчим
NOT_NODES = {
    "laneSet", "sequenceFlow", "dataObject", "dataObjectReference",
    "documentation", "extensionElements", "property", "ioSpecification",
    "dataInput", "dataOutput", "incoming", "outgoing",
    "dataInputAssociation", "dataOutputAssociation", "standardLoopCharacteristics",
    "multiInstanceLoopCharacteristics",
}
# Несут смысл, но в таблице места нет — о них предупреждаем
LOST_TAGS = {
    "textAnnotation": "текстовая аннотация",
    "association": "связь с аннотацией",
    "group": "рамка-группа",
    "dataStoreReference": "хранилище данных",
}

# Символы, ломающие разбор ячеек
CELL_BAD = "|"
FLOW_BAD = str.maketrans({"[": "(", "]": ")", "{": "(", "}": ")", ";": ","})


@dataclass
class TableResult:
    ok: bool
    table: str | None
    issues: list[Issue]
    stats: dict[str, int]
    rows: list[dict[str, str]] = field(default_factory=list)
    columns: list[str] = field(default_factory=list)

    @property
    def warnings(self) -> list[Issue]:
        return [i for i in self.issues if i.level == "warning"]

    @property
    def errors(self) -> list[Issue]:
        return [i for i in self.issues if i.level == "error"]

    def as_dict(self) -> dict[str, Any]:
        return {"ok": self.ok, "table": self.table, "columns": self.columns,
                "issues": [i.as_dict() for i in self.issues], "stats": self.stats}


# --------------------------------------------------------------------------- #
# Утилиты
# --------------------------------------------------------------------------- #

def _tag(el: ET.Element) -> str:
    return el.tag.split("}")[-1]


def _text(el: ET.Element | None) -> str:
    return " ".join((el.text or "").split()) if el is not None else ""


def _clean(value: str, issues: list[Issue], where: str) -> str:
    """Схлопывает переводы строк и убирает символ, ломающий формат."""
    text = " ".join(str(value or "").split())
    if CELL_BAD in text:
        text = text.replace(CELL_BAD, "/")
        issues.append(Issue("warning", None,
                            f"{where}: символ '|' заменён на '/', иначе таблица "
                            f"не читается"))
    return text


def _clean_flow(value: str, issues: list[Issue], where: str) -> str:
    """То же, но ещё и для скобок, которые в колонке next служебные."""
    text = _clean(value, issues, where)
    fixed = text.translate(FLOW_BAD)
    if fixed != text:
        issues.append(Issue("warning", None,
                            f"{where}: скобки и ';' заменены — в колонке next "
                            f"они служебные"))
    return fixed


def _parse(xml: str | bytes) -> ET.Element:
    try:
        if isinstance(xml, bytes):
            for encoding in ("utf-8-sig", "utf-8", "cp1251"):
                try:
                    xml = xml.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ConvertError("Не удалось определить кодировку файла")
        return ET.fromstring(xml)
    except ET.ParseError as exc:
        raise ConvertError(f"Не удалось разобрать XML: {exc}") from exc


# --------------------------------------------------------------------------- #
# Разбор
# --------------------------------------------------------------------------- #

def _lane_map(proc: ET.Element, issues: list[Issue]) -> dict[str, str]:
    """id узла -> название дорожки. Вложенные дорожки схлопываются."""
    out: dict[str, str] = {}
    for lane in proc.iter(f"{B}lane"):
        name = lane.get("name") or lane.get("id", "")
        for ref in lane.findall(f"{B}flowNodeRef"):
            nid = (ref.text or "").strip()
            if not nid:
                continue
            if nid in out and out[nid] != name:
                issues.append(Issue("warning", None,
                                    f"{nid!r} состоит сразу в дорожках {out[nid]!r} "
                                    f"и {name!r}; оставлена вторая"))
            out[nid] = name
    return out


def _doc_map(proc: ET.Element) -> dict[str, str]:
    """id ссылки на документ -> id самого документа (он же id строки таблицы).

    Ищем по всему дереву: документы объявляют и внутри субпроцессов.
    """
    out: dict[str, str] = {}
    for ref in proc.iter(f"{B}dataObjectReference"):
        rid = ref.get("id", "")
        out[rid] = ref.get("dataObjectRef") or rid
    return out


def _node_row(el: ET.Element, pool: str, lane: str, parent: str,
              docs: dict[str, str], issues: list[Issue]) -> dict[str, str]:
    nid = el.get("id", "")
    row = {c: "" for c in COLUMN_ORDER}
    row["id"] = nid
    row["name"] = _clean(el.get("name", ""), issues, f"название {nid!r}")
    row["type"] = _tag(el)
    row["pool"] = _clean(pool, issues, "название пула")
    row["lane"] = _clean(lane, issues, "название дорожки")
    row["parent_id"] = parent
    row["documentation"] = _clean(_text(el.find(f"{B}documentation")),
                                  issues, f"описание {nid!r}")

    if row["type"] == "boundaryEvent":
        row["attached_to"] = el.get("attachedToRef", "")
        if el.get("cancelActivity", "true").lower() == "false":
            row["cancel_activity"] = "FALSE"

    data_in, data_out = [], []
    for child in el:
        tag = _tag(child)
        if tag in EVENT_DEF_BY_TAG:
            if row["event_def"]:
                issues.append(Issue("warning", None,
                                    f"у события {nid!r} несколько определений; "
                                    f"в таблицу попало {row['event_def']!r}"))
            else:
                row["event_def"] = EVENT_DEF_BY_TAG[tag]
        elif tag == "standardLoopCharacteristics":
            row["marker"] = "loop"
        elif tag == "multiInstanceLoopCharacteristics":
            row["marker"] = ("mi_sequential"
                             if child.get("isSequential", "false").lower() == "true"
                             else "mi_parallel")
        elif tag == "dataInputAssociation":
            ref = _text(child.find(f"{B}sourceRef"))
            if ref:
                data_in.append(docs.get(ref, ref))
        elif tag == "dataOutputAssociation":
            ref = _text(child.find(f"{B}targetRef"))
            if ref:
                data_out.append(docs.get(ref, ref))
        elif tag == "extensionElements" and len(child):
            issues.append(Issue("warning", None,
                                f"у {nid!r} есть расширения движка "
                                f"({', '.join(sorted({_tag(g) for g in child}))}); "
                                f"в таблицу они не переносятся"))

    row["data_in"] = ";".join(data_in)
    row["data_out"] = ";".join(data_out)
    return row


def _walk(container: ET.Element, pool: str, parent: str, lanes: dict[str, str],
          docs: dict[str, str], rows: list[dict[str, str]],
          defaults: set[str], issues: list[Issue]) -> None:
    """Собирает узлы контейнера, спускаясь в субпроцессы."""
    for el in container:
        tag = _tag(el)
        if tag in NOT_NODES:
            continue
        if tag in LOST_TAGS:
            issues.append(Issue("warning", None,
                                f"{LOST_TAGS[tag]} {el.get('id', '')!r} "
                                f"в таблице не представима, потеряна"))
            continue
        if tag not in NODE_TYPES:
            issues.append(Issue("warning", None,
                                f"элемент {el.get('id', '')!r} типа {tag!r} "
                                f"таблицей не поддерживается, пропущен"))
            continue

        if el.get("default"):
            defaults.add(el.get("default", ""))
        rows.append(_node_row(el, pool, lanes.get(el.get("id", ""), ""),
                              parent, docs, issues))
        if tag in SUBPROCESS_TYPES:
            _walk(el, pool, el.get("id", ""), lanes, docs, rows, defaults, issues)


def _flow_item(flow: ET.Element, defaults: set[str],
               issues: list[Issue]) -> tuple[str, str, str]:
    """sequenceFlow -> (id источника, id цели, запись для колонки next)."""
    fid = flow.get("id", "")
    target = flow.get("targetRef", "")
    item = f"*{target}" if fid in defaults else target

    name = _clean_flow(flow.get("name", ""), issues, f"подпись потока {fid!r}")
    if name:
        item += f" [{name}]"
    condition = _clean_flow(_text(flow.find(f"{B}conditionExpression")),
                            issues, f"условие потока {fid!r}")
    if condition:
        item += f" {{{condition}}}"
    return flow.get("sourceRef", ""), target, item


def _kept(refs: list[str], known: set[str], issues: list[Issue],
          where: str) -> list[str]:
    """Выбрасывает ссылки на элементы, которых в таблице нет."""
    out: list[str] = []
    for ref in refs:
        if ref in known:
            out.append(ref)
        else:
            issues.append(Issue("warning", None,
                                f"{where}: ссылка на {ref!r}, которого в таблице "
                                f"нет; убрана, иначе таблица не соберётся"))
    return out


def _link(all_rows: list[dict[str, str]], next_by_src: dict[str, list[tuple[str, str]]],
          msg_flows: list[tuple[str, str]], issues: list[Issue]) -> None:
    """Расставляет потоки и сообщения, отсекая всё, что повисло в воздухе."""
    by_id = {r["id"]: r for r in all_rows}
    known = set(by_id)

    for src, pairs in next_by_src.items():
        if src not in by_id:
            issues.append(Issue("warning", None,
                                f"поток исходит из {src!r}, которого нет среди "
                                f"элементов; потерян"))
            continue
        items = [item for target, item in pairs
                 if _kept([target], known, issues, f"поток из {src!r}")]
        by_id[src]["next"] = "; ".join(items)

    for src, tgt in msg_flows:
        if src not in by_id:
            issues.append(Issue("warning", None,
                                f"поток сообщений из {src!r} не привязан к элементу "
                                f"(вероятно, идёт от пула-ящика), потерян"))
            continue
        if not _kept([tgt], known, issues, f"сообщение из {src!r}"):
            continue
        current = by_id[src]["message_to"]
        by_id[src]["message_to"] = f"{current};{tgt}" if current else tgt

    for row in all_rows:
        where = f"элемент {row['id']!r}"
        for column in ("data_in", "data_out"):
            if row[column]:
                row[column] = ";".join(_kept(row[column].split(";"), known,
                                             issues, f"{where}, {column}"))
        for column in ("attached_to", "parent_id"):
            if row[column] and not _kept([row[column]], known, issues,
                                         f"{where}, {column}"):
                row[column] = ""


# --------------------------------------------------------------------------- #
# Точка входа
# --------------------------------------------------------------------------- #

def to_table(xml: str | bytes, *, full: bool = False,
             sep: str = "|") -> TableResult:
    """Разбирает BPMN в таблицу. Ничего не читает и не пишет на диск."""
    issues: list[Issue] = []
    root = _parse(xml)

    processes = root.findall(f"{B}process")
    if not processes:
        raise ConvertError("В документе нет ни одного process")

    # --- участники ---------------------------------------------------------
    pools: dict[str, str] = {}
    msg_flows: list[tuple[str, str]] = []
    for collab in root.findall(f"{B}collaboration"):
        for part in collab.findall(f"{B}participant"):
            ref = part.get("processRef")
            name = part.get("name") or part.get("id", "")
            if not ref:
                issues.append(Issue("warning", None,
                                    f"пул {name!r} без процесса (чёрный ящик) "
                                    f"в таблице не представим, пропущен"))
                continue
            pools[ref] = name
        for mf in collab.findall(f"{B}messageFlow"):
            msg_flows.append((mf.get("sourceRef", ""), mf.get("targetRef", "")))

    # --- узлы --------------------------------------------------------------
    rows: list[dict[str, str]] = []
    doc_rows: list[dict[str, str]] = []
    defaults: set[str] = set()
    next_by_src: dict[str, list[str]] = {}
    flow_count = 0

    for proc in processes:
        pid = proc.get("id", "")
        pool = pools.get(pid) or proc.get("name") or pid
        lanes = _lane_map(proc, issues)
        docs = _doc_map(proc)

        _walk(proc, pool, "", lanes, docs, rows, defaults, issues)

        seen: set[str] = set()
        for ref in proc.iter(f"{B}dataObjectReference"):
            table_id = docs.get(ref.get("id", ""), ref.get("id", ""))
            if table_id in seen:
                issues.append(Issue("warning", None,
                                    f"на документ {table_id!r} несколько ссылок; "
                                    f"в таблице он будет один"))
                continue
            seen.add(table_id)
            obj = next((o for o in proc.iter(f"{B}dataObject")
                        if o.get("id") == table_id), None)
            row = {c: "" for c in COLUMN_ORDER}
            row.update(id=table_id, type="dataObject",
                       name=_clean(ref.get("name", ""), issues,
                                   f"название документа {table_id!r}"),
                       pool=_clean(pool, issues, "название пула"),
                       documentation=_clean(_text(obj.find(f"{B}documentation"))
                                            if obj is not None else "",
                                            issues, f"описание {table_id!r}"))
            doc_rows.append(row)

        for flow in proc.iter(f"{B}sequenceFlow"):
            flow_count += 1
            src, target, item = _flow_item(flow, defaults, issues)
            next_by_src.setdefault(src, []).append((target, item))

    all_rows = rows + doc_rows
    if not all_rows:
        raise ConvertError("В документе не нашлось ни одного элемента")

    _link(all_rows, next_by_src, msg_flows, issues)

    # --- какие колонки выводить -------------------------------------------
    if full:
        columns = list(COLUMN_ORDER)
    else:
        columns = [c for c in COLUMN_ORDER
                   if c in ALWAYS or any(r[c] for r in all_rows)]

    lines = [sep.join(columns)]
    lines += [sep.join(r[c] for c in columns) for r in all_rows]

    stats = {
        "nodes": len(all_rows),
        "flows": flow_count,
        "message_flows": len(msg_flows),
        "pools": len({r["pool"] for r in all_rows}),
        "columns": len(columns),
        "warnings": sum(1 for i in issues if i.level == "warning"),
    }
    return TableResult(ok=True, table="\n".join(lines), issues=issues,
                       stats=stats, rows=all_rows, columns=columns)


def to_workbook(result: TableResult, template: bytes) -> bytes:
    """Раскладывает строки по вложенному шаблону. Возвращает байты .xlsx.

    Шаблон нужен ради выпадающих списков и листа «Справка»: заполненная
    таблица остаётся редактируемой ровно так же, как пустая.
    """
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(template))
    ws = wb["Process"]
    if ws.max_row > 1:                      # выкинуть строки-примеры вместе с их заливкой
        ws.delete_rows(2, ws.max_row - 1)
    for row in result.rows:
        ws.append([row[column] or None for column in COLUMN_ORDER])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
