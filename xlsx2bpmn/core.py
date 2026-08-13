# -*- coding: utf-8 -*-
"""
Ядро конвертера: таблица (xlsx) -> BPMN 2.0 XML.

Чистая логика без ввода-вывода и без сети.
Единственная точка входа: convert(data: bytes, ...) -> Result
"""
from __future__ import annotations

import io
import re
import unicodedata
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

# --------------------------------------------------------------------------- #
# Константы и словари
# --------------------------------------------------------------------------- #

MAX_ROWS = 10_000

NS = {
    "bpmn": "http://www.omg.org/spec/BPMN/20100524/MODEL",
    "bpmndi": "http://www.omg.org/spec/BPMN/20100524/DI",
    "dc": "http://www.omg.org/spec/DD/20100524/DC",
    "di": "http://www.omg.org/spec/DD/20100524/DI",
    "xsi": "http://www.w3.org/2001/XMLSchema-instance",
    "zeebe": "http://camunda.org/schema/zeebe/1.0",
    "modeler": "http://camunda.org/schema/modeler/1.0",
}
CAMUNDA_VERSION = "8.9.0"
JOB_TASK_TYPES = {"serviceTask", "sendTask", "scriptTask", "businessRuleTask"}
TARGET_NS = "http://bpmn.io/schema/bpmn"

EVENT_TYPES = {
    "startEvent", "endEvent", "intermediateCatchEvent",
    "intermediateThrowEvent", "boundaryEvent",
}
TASK_TYPES = {
    "task", "userTask", "serviceTask", "scriptTask", "manualTask",
    "sendTask", "receiveTask", "businessRuleTask",
}
SUBPROCESS_TYPES = {"subProcess", "callActivity"}
ACTIVITY_TYPES = TASK_TYPES | SUBPROCESS_TYPES
DATA_TYPES = {"dataObject"}
GATEWAY_TYPES = {
    "exclusiveGateway", "parallelGateway",
    "inclusiveGateway", "eventBasedGateway",
}
NODE_TYPES = EVENT_TYPES | ACTIVITY_TYPES | GATEWAY_TYPES | DATA_TYPES

EVENT_DEFS = {
    "message": "messageEventDefinition",
    "timer": "timerEventDefinition",
    "error": "errorEventDefinition",
    "signal": "signalEventDefinition",
    "escalation": "escalationEventDefinition",
    "conditional": "conditionalEventDefinition",
    "terminate": "terminateEventDefinition",
    "link": "linkEventDefinition",
    "compensation": "compensateEventDefinition",
}

MARKERS = {"loop", "mi_parallel", "mi_sequential"}

# Заголовки: нормализованное имя -> каноническое
COLUMNS = {
    "id": "id",
    "name": "name",
    "type": "type",
    "eventdef": "event_def",
    "eventdefinition": "event_def",
    "pool": "pool",
    "lane": "lane",
    "parentid": "parent_id",
    "attachedto": "attached_to",
    "cancelactivity": "cancel_activity",
    "marker": "marker",
    "next": "next",
    "messageto": "message_to",
    "documentation": "documentation",
    "datain": "data_in",
    "dataout": "data_out",
}
REQUIRED_COLUMNS = {"id", "type", "pool"}

TRUE_WORDS = {"true", "1", "yes", "y", "да", "истина", "+"}
FALSE_WORDS = {"false", "0", "no", "n", "нет", "ложь", "-"}

FLOW_ITEM_RE = re.compile(r"\[([^\]]*)\]|\{([^}]*)\}")
NCNAME_BAD = re.compile(r"[^A-Za-z0-9_.\-]")


# --------------------------------------------------------------------------- #
# Модель
# --------------------------------------------------------------------------- #

@dataclass
class Issue:
    level: str          # "error" | "warning"
    row: int | None
    message: str

    def as_dict(self) -> dict[str, Any]:
        return {"level": self.level, "row": self.row, "message": self.message}

    def __str__(self) -> str:
        where = f"строка {self.row}" if self.row else "файл"
        tag = "ERROR" if self.level == "error" else "WARN "
        return f"[{where}] {tag} {self.message}"


@dataclass
class Flow:
    id: str
    source: str
    target: str
    name: str = ""
    condition: str = ""
    is_default: bool = False
    row: int | None = None


@dataclass
class MessageFlow:
    id: str
    source: str
    target: str
    row: int | None = None


@dataclass
class Node:
    id: str
    name: str
    type: str
    event_def: str = ""
    pool: str = ""
    lane: str = ""
    parent_id: str = ""
    attached_to: str = ""
    cancel_activity: bool = True
    marker: str = ""
    documentation: str = ""
    data_in: list[str] = field(default_factory=list)
    data_out: list[str] = field(default_factory=list)
    row: int | None = None
    incoming: list[str] = field(default_factory=list)
    outgoing: list[str] = field(default_factory=list)


@dataclass
class Result:
    ok: bool
    xml: str | None
    issues: list[Issue]
    stats: dict[str, int]

    @property
    def errors(self) -> list[Issue]:
        return [i for i in self.issues if i.level == "error"]

    @property
    def warnings(self) -> list[Issue]:
        return [i for i in self.issues if i.level == "warning"]

    def as_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "xml": self.xml,
            "issues": [i.as_dict() for i in self.issues],
            "stats": self.stats,
        }


class ConvertError(Exception):
    """Вход невозможно прочитать вообще."""


# --------------------------------------------------------------------------- #
# Утилиты
# --------------------------------------------------------------------------- #

def _s(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_header(value: Any) -> str:
    text = _s(value).lower()
    return re.sub(r"[^a-z0-9]", "", text)


def _tribool(value: str, default: bool = True) -> bool:
    v = value.strip().lower()
    if v in TRUE_WORDS:
        return True
    if v in FALSE_WORDS:
        return False
    return default


def _sanitize_id(raw: str) -> str:
    """Приводит идентификатор к валидному XML NCName."""
    text = unicodedata.normalize("NFKD", raw)
    text = NCNAME_BAD.sub("_", text)
    if not text or not re.match(r"[A-Za-z_]", text[0]):
        text = "n_" + text
    return text


def parse_next(raw: str) -> list[dict[str, Any]]:
    """'*t_a [Нет]; t_b [Да] {x > 1}' -> список описаний потоков."""
    out: list[dict[str, Any]] = []
    for chunk in raw.split(";"):
        item = chunk.strip()
        if not item:
            continue
        label, condition = "", ""
        for match in FLOW_ITEM_RE.finditer(item):
            if match.group(1) is not None:
                label = match.group(1).strip()
            else:
                condition = (match.group(2) or "").strip()
        rest = FLOW_ITEM_RE.sub("", item).strip()
        is_default = rest.startswith("*")
        target = rest.lstrip("*").strip()
        out.append({
            "target": target,
            "name": label,
            "condition": condition,
            "is_default": is_default,
            "raw": item,
        })
    return out


# --------------------------------------------------------------------------- #
# Чтение таблицы
# --------------------------------------------------------------------------- #

def _text_rows(data: bytes, issues: list[Issue]):
    """Читает текстовую таблицу. Разделитель только | — другие не принимаются."""
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ConvertError("Не удалось определить кодировку файла")

    lines = [l for l in text.split("\n") if l.strip()]
    if not lines:
        raise ConvertError("Файл пустой")
    if "|" not in lines[0]:
        raise ConvertError(
            "В текстовом файле не найден разделитель '|'. Колонки разделяются "
            "вертикальной чертой; табуляции, запятые и точки с запятой "
            "разделителями не считаются")

    width = lines[0].count("|") + 1
    rows = []
    for number, line in enumerate(lines, start=1):
        cells = line.split("|")
        if len(cells) < width:
            issues.append(Issue("warning", number,
                                f"ячеек {len(cells)}, ожидается {width}; "
                                f"недостающие считаю пустыми"))
            cells += [""] * (width - len(cells))
        elif len(cells) > width:
            issues.append(Issue("error", number,
                                f"ячеек {len(cells)}, ожидается {width}; "
                                f"лишняя черта '|' сдвигает колонки"))
            continue
        rows.append(tuple(cells))
    return rows


def _sheet_rows(data: bytes, sheet: str | None, issues: list[Issue]):
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:                       # noqa: BLE001
        raise ConvertError(f"Не удалось прочитать xlsx: {exc}") from exc
    try:
        if sheet and sheet in wb.sheetnames:
            ws = wb[sheet]
        elif "Process" in wb.sheetnames:
            ws = wb["Process"]
        else:
            ws = wb[wb.sheetnames[0]]
            issues.append(Issue("warning", None,
                                f"Лист 'Process' не найден, читаю '{ws.title}'"))
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def read_table(data: bytes, sheet: str | None = None) -> tuple[list[Node], list[Flow],
                                                               list[MessageFlow], list[Issue]]:
    issues: list[Issue] = []
    # xlsx — это zip, начинается с PK; всё остальное считаем текстом
    all_rows = (_sheet_rows(data, sheet, issues) if data[:2] == b"PK"
                else _text_rows(data, issues))

    rows = iter(all_rows)
    try:
        header = next(rows)
    except StopIteration:
        raise ConvertError("Файл пустой")

    col_map: dict[str, int] = {}
    for idx, cell in enumerate(header):
        key = COLUMNS.get(_norm_header(cell))
        if key and key not in col_map:
            col_map[key] = idx

    missing = REQUIRED_COLUMNS - col_map.keys()
    if missing:
        raise ConvertError(
            "В таблице нет обязательных колонок: " + ", ".join(sorted(missing))
        )

    def cell(row: tuple, key: str) -> str:
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        return _s(row[idx])

    nodes: list[Node] = []
    flows: list[Flow] = []
    msgs: list[MessageFlow] = []
    flow_ids: dict[str, int] = {}

    for offset, row in enumerate(rows, start=2):
        if offset > MAX_ROWS:
            issues.append(Issue("error", None,
                                f"Превышен лимит строк ({MAX_ROWS})"))
            break
        if row is None or not any(_s(v) for v in row):
            continue

        node = Node(
            id=cell(row, "id"),
            name=cell(row, "name"),
            type=cell(row, "type"),
            event_def=("" if cell(row, "event_def").lower() == "none"
                       else cell(row, "event_def").lower()),
            pool=cell(row, "pool"),
            lane=cell(row, "lane"),
            parent_id=cell(row, "parent_id"),
            attached_to=cell(row, "attached_to"),
            cancel_activity=_tribool(cell(row, "cancel_activity")),
            marker=cell(row, "marker").lower(),
            documentation=cell(row, "documentation"),
            data_in=[d.strip() for d in cell(row, "data_in").split(";") if d.strip()],
            data_out=[d.strip() for d in cell(row, "data_out").split(";") if d.strip()],
            row=offset,
        )
        if not node.id:
            issues.append(Issue("error", offset, "не заполнена колонка id"))
            continue
        nodes.append(node)

        for spec in parse_next(cell(row, "next")):
            if not spec["target"]:
                issues.append(Issue("error", offset,
                                    f"не разобран поток: {spec['raw']!r}"))
                continue
            base = f"f_{_sanitize_id(node.id)}_{_sanitize_id(spec['target'])}"
            flow_ids[base] = flow_ids.get(base, 0) + 1
            fid = base if flow_ids[base] == 1 else f"{base}_{flow_ids[base]}"
            flows.append(Flow(
                id=fid, source=node.id, target=spec["target"],
                name=spec["name"], condition=spec["condition"],
                is_default=spec["is_default"], row=offset,
            ))

        for chunk in cell(row, "message_to").split(";"):
            target = chunk.strip()
            if target:
                msgs.append(MessageFlow(
                    id=f"mf_{_sanitize_id(node.id)}_{_sanitize_id(target)}",
                    source=node.id, target=target, row=offset,
                ))
    return nodes, flows, msgs, issues


# --------------------------------------------------------------------------- #
# Валидация
# --------------------------------------------------------------------------- #

def validate(nodes: list[Node], flows: list[Flow],
             msgs: list[MessageFlow]) -> tuple[dict[str, Node], list[Issue]]:
    issues: list[Issue] = []
    by_id: dict[str, Node] = {}

    for node in nodes:
        if node.id in by_id:
            issues.append(Issue("error", node.row,
                                f"дубль id {node.id!r} (уже был в строке {by_id[node.id].row})"))
            continue
        by_id[node.id] = node

    for node in by_id.values():
        if node.type not in NODE_TYPES:
            issues.append(Issue("error", node.row,
                                f"неизвестный type {node.type!r}"))
        if not node.pool:
            issues.append(Issue("error", node.row, "не заполнена колонка pool"))
        if not node.name and node.type not in GATEWAY_TYPES:
            issues.append(Issue("warning", node.row, f"у {node.id!r} пустое name"))
        if node.event_def:
            if node.event_def not in EVENT_DEFS:
                issues.append(Issue("error", node.row,
                                    f"неизвестный event_def {node.event_def!r}"))
            elif node.type not in EVENT_TYPES:
                issues.append(Issue("warning", node.row,
                                    f"event_def указан у не-события {node.id!r}, игнорирую"))
        if node.marker and node.marker not in MARKERS:
            issues.append(Issue("error", node.row, f"неизвестный marker {node.marker!r}"))
        if node.marker and node.type not in ACTIVITY_TYPES:
            issues.append(Issue("warning", node.row,
                                f"marker применим только к активностям, у {node.id!r} игнорирую"))

        for direction, refs in (("data_in", node.data_in), ("data_out", node.data_out)):
            if refs and node.type not in ACTIVITY_TYPES:
                issues.append(Issue("error", node.row,
                                    f"{direction} задан у {node.type}; документы крепятся "
                                    f"только к задачам и субпроцессам"))
                continue
            for ref in refs:
                doc = by_id.get(ref)
                if doc is None:
                    issues.append(Issue("error", node.row,
                                        f"{direction}: документ {ref!r} не найден"))
                elif doc.type not in DATA_TYPES:
                    issues.append(Issue("error", node.row,
                                        f"{direction}: {ref!r} — это {doc.type}, а не документ"))
                elif doc.pool != node.pool:
                    issues.append(Issue("error", node.row,
                                        f"{direction}: документ {ref!r} объявлен в другом пуле; "
                                        f"между пулами документы передаются потоком сообщений"))

        if node.parent_id:
            parent = by_id.get(node.parent_id)
            if parent is None:
                issues.append(Issue("error", node.row,
                                    f"parent_id {node.parent_id!r} не найден"))
            elif parent.type not in SUBPROCESS_TYPES:
                issues.append(Issue("error", node.row,
                                    f"parent_id {node.parent_id!r} не является субпроцессом"))
            elif parent.pool != node.pool:
                issues.append(Issue("error", node.row,
                                    f"{node.id!r} и его субпроцесс лежат в разных пулах"))

        if node.type == "boundaryEvent":
            if not node.attached_to:
                issues.append(Issue("error", node.row,
                                    f"у boundaryEvent {node.id!r} не заполнен attached_to"))
            else:
                host = by_id.get(node.attached_to)
                if host is None:
                    issues.append(Issue("error", node.row,
                                        f"attached_to {node.attached_to!r} не найден"))
                elif host.type not in ACTIVITY_TYPES:
                    issues.append(Issue("error", node.row,
                                        f"boundaryEvent можно крепить только к активности, "
                                        f"а {node.attached_to!r} — {host.type}"))
                elif host.pool != node.pool:
                    issues.append(Issue("error", node.row,
                                        f"boundaryEvent и его активность в разных пулах"))
        elif node.attached_to:
            issues.append(Issue("warning", node.row,
                                f"attached_to задан у {node.type}, игнорирую"))

    # потоки
    live_flows: list[Flow] = []
    for flow in flows:
        src, tgt = by_id.get(flow.source), by_id.get(flow.target)
        if (src is not None and src.type in DATA_TYPES) or (tgt is not None and tgt.type in DATA_TYPES):
            issues.append(Issue("error", flow.row,
                                "поток управления не может вести к документу; "
                                "используйте data_in / data_out"))
            continue
        if tgt is None:
            issues.append(Issue("error", flow.row,
                                f"next ссылается на несуществующий id {flow.target!r}"))
            continue
        if src is None:
            continue
        if src.pool != tgt.pool:
            issues.append(Issue("error", flow.row,
                                f"поток {flow.source!r} -> {flow.target!r} пересекает пулы; "
                                f"между пулами используйте message_to"))
            continue
        if (src.parent_id or "") != (tgt.parent_id or "") and tgt.type != "boundaryEvent":
            issues.append(Issue("warning", flow.row,
                                f"поток {flow.source!r} -> {flow.target!r} пересекает границу "
                                f"субпроцесса; помещаю на верхний уровень"))
        if src.type == "endEvent":
            issues.append(Issue("error", flow.row,
                                f"у endEvent {src.id!r} не может быть исходящих потоков"))
            continue
        if tgt.type == "startEvent":
            issues.append(Issue("error", flow.row,
                                f"в startEvent {tgt.id!r} не может входить поток"))
            continue
        if tgt.type == "boundaryEvent":
            issues.append(Issue("error", flow.row,
                                f"в boundaryEvent {tgt.id!r} не может входить поток"))
            continue
        src.outgoing.append(flow.id)
        tgt.incoming.append(flow.id)
        live_flows.append(flow)

    # boundary -> outgoing хоста не считается входящим потоком
    used_docs = {r for n in by_id.values() for r in (n.data_in + n.data_out)}
    for node in by_id.values():
        if node.type in DATA_TYPES:
            if node.id not in used_docs:
                issues.append(Issue("warning", node.row,
                                    f"документ {node.id!r} ни одной задачей не используется"))
            continue
        if node.type == "boundaryEvent":
            continue
        if node.type != "startEvent" and not node.incoming:
            issues.append(Issue("warning", node.row,
                                f"у {node.id!r} нет входящих потоков"))
        if node.type != "endEvent" and not node.outgoing:
            issues.append(Issue("warning", node.row,
                                f"у {node.id!r} нет исходящих потоков"))

    # шлюзы
    flows_by_id = {f.id: f for f in live_flows}
    for node in by_id.values():
        if node.type not in {"exclusiveGateway", "inclusiveGateway"}:
            continue
        outs = [flows_by_id[i] for i in node.outgoing]
        if len(outs) < 2:
            continue
        defaults = [f for f in outs if f.is_default]
        conditioned = [f for f in outs if f.condition]
        if len(defaults) > 1:
            issues.append(Issue("error", node.row,
                                f"у шлюза {node.id!r} несколько потоков по умолчанию"))
        if not defaults and len(conditioned) < len(outs):
            issues.append(Issue("error", node.row,
                                f"у шлюза {node.id!r} {len(outs)} исходящих потока, "
                                f"условия заданы у {len(conditioned)}, "
                                f"поток по умолчанию не указан"))
        for flow in defaults:
            if flow.condition:
                issues.append(Issue("warning", flow.row,
                                    f"у потока по умолчанию {flow.id!r} задано условие, игнорирую"))
                flow.condition = ""

    for node in by_id.values():
        if node.type == "parallelGateway":
            for fid in node.outgoing:
                if flows_by_id[fid].condition:
                    issues.append(Issue("warning", flows_by_id[fid].row,
                                        f"условие на исходящем из parallelGateway {node.id!r} "
                                        f"игнорируется BPMN-движками"))

    # потоки сообщений
    live_msgs: list[MessageFlow] = []
    for msg in msgs:
        src, tgt = by_id.get(msg.source), by_id.get(msg.target)
        if tgt is None:
            issues.append(Issue("error", msg.row,
                                f"message_to ссылается на несуществующий id {msg.target!r}"))
            continue
        if src is None:
            continue
        if src.pool == tgt.pool:
            issues.append(Issue("error", msg.row,
                                f"message_to {msg.target!r} ведёт в тот же пул; "
                                f"внутри пула используйте next"))
            continue
        live_msgs.append(msg)

    # пулы
    pools: dict[str, list[Node]] = {}
    for node in by_id.values():
        if node.type not in DATA_TYPES:
            pools.setdefault(node.pool, []).append(node)
    for pool, members in pools.items():
        if not any(n.type == "startEvent" for n in members):
            issues.append(Issue("error", None, f"в пуле {pool!r} нет startEvent"))
        if not any(n.type == "endEvent" for n in members):
            issues.append(Issue("error", None, f"в пуле {pool!r} нет endEvent"))

    return by_id, issues


# --------------------------------------------------------------------------- #
# Генерация XML
# --------------------------------------------------------------------------- #

def _q(prefix: str, tag: str) -> str:
    return f"{{{NS[prefix]}}}{tag}"


def build_xml(by_id: dict[str, Node], flows: list[Flow], msgs: list[MessageFlow],
              executable: bool = False, target: str = "plain") -> str:
    ids: dict[str, str] = {}
    used: set[str] = set()
    for key in by_id:
        candidate = _sanitize_id(key)
        base, n = candidate, 1
        while candidate in used:
            n += 1
            candidate = f"{base}_{n}"
        used.add(candidate)
        ids[key] = candidate

    pools: dict[str, list[Node]] = {}
    for node in by_id.values():
        pools.setdefault(node.pool, []).append(node)

    pool_proc: dict[str, str] = {}
    for i, pool in enumerate(pools, start=1):
        pool_proc[pool] = f"Process_{i}"

    for prefix, uri in NS.items():
        ET.register_namespace(prefix, uri)

    root_attrs = {
        "id": "Definitions_1",
        "targetNamespace": TARGET_NS,
        "exporter": "xlsx2bpmn",
        "exporterVersion": "1.0",
    }
    if target == "camunda8":
        root_attrs[_q("modeler", "executionPlatform")] = "Camunda Cloud"
        root_attrs[_q("modeler", "executionPlatformVersion")] = CAMUNDA_VERSION
    defs = ET.Element(_q("bpmn", "definitions"), root_attrs)

    # рамка пула нужна и одиночному процессу, если в нём есть дорожки:
    # без участника дорожки повисают в воздухе, редакторы их так не ждут
    multi = (len(pools) > 1 or bool(msgs)
             or any(n.lane for n in by_id.values() if n.type not in DATA_TYPES))
    collab_id = "Collaboration_1"
    if multi:
        collab = ET.SubElement(defs, _q("bpmn", "collaboration"), {"id": collab_id})
        for i, pool in enumerate(pools, start=1):
            ET.SubElement(collab, _q("bpmn", "participant"), {
                "id": f"Participant_{i}", "name": pool, "processRef": pool_proc[pool],
            })
        for msg in msgs:
            ET.SubElement(collab, _q("bpmn", "messageFlow"), {
                "id": msg.id,
                "sourceRef": ids[msg.source],
                "targetRef": ids[msg.target],
            })

    doc_ref: dict[str, str] = {}
    for node in by_id.values():
        if node.type in DATA_TYPES:
            doc_ref[node.id] = f"{ids[node.id]}_ref"

    flows_by_src: dict[str, list[Flow]] = {}
    for flow in flows:
        flows_by_src.setdefault(flow.source, []).append(flow)

    for pool, members in pools.items():
        proc = ET.SubElement(defs, _q("bpmn", "process"), {
            "id": pool_proc[pool],
            "name": pool,
            "isExecutable": "true" if executable else "false",
        })

        lanes: dict[str, list[Node]] = {}
        for node in members:
            if node.lane and not node.parent_id and node.type not in DATA_TYPES:
                lanes.setdefault(node.lane, []).append(node)
        if lanes:
            lane_set = ET.SubElement(proc, _q("bpmn", "laneSet"),
                                     {"id": f"LaneSet_{pool_proc[pool]}"})
            for i, (lane_name, lane_nodes) in enumerate(lanes.items(), start=1):
                lane = ET.SubElement(lane_set, _q("bpmn", "lane"), {
                    "id": f"Lane_{pool_proc[pool]}_{i}", "name": lane_name,
                })
                for node in lane_nodes:
                    ref = ET.SubElement(lane, _q("bpmn", "flowNodeRef"))
                    ref.text = ids[node.id]

        docs = [n for n in members if n.type in DATA_TYPES]
        top = [n for n in members if not n.parent_id and n.type not in DATA_TYPES]
        children: dict[str, list[Node]] = {}
        for node in members:
            if node.parent_id:
                children.setdefault(node.parent_id, []).append(node)

        def emit(container: ET.Element, node_list: list[Node]) -> None:
            for node in node_list:
                el = _emit_node(container, node, ids, target, doc_ref)
                if node.type in SUBPROCESS_TYPES and node.id in children:
                    emit(el, children[node.id])
                    for child in children[node.id]:
                        for flow in flows_by_src.get(child.id, []):
                            if flow.id in placed:
                                continue
                            _emit_flow(el, flow, ids)
                            placed.add(flow.id)

        for doc in docs:
            obj = ET.SubElement(proc, _q("bpmn", "dataObject"), {"id": ids[doc.id]})
            if doc.documentation:
                ET.SubElement(obj, _q("bpmn", "documentation")).text = doc.documentation
            ET.SubElement(proc, _q("bpmn", "dataObjectReference"), {
                "id": doc_ref[doc.id], "name": doc.name or doc.id,
                "dataObjectRef": ids[doc.id],
            })

        placed: set[str] = set()
        emit(proc, top)
        for flow in flows:
            if flow.id in placed or by_id[flow.source].pool != pool:
                continue
            _emit_flow(proc, flow, ids)
            placed.add(flow.id)

    ET.indent(defs, space="  ")
    return ET.tostring(defs, encoding="unicode", xml_declaration=True)


def _emit_node(parent: ET.Element, node: Node, ids: dict[str, str],
               target: str = "plain",
               doc_ref: dict[str, str] | None = None) -> ET.Element:
    attrs = {"id": ids[node.id]}
    if node.name:
        attrs["name"] = node.name
    if node.type == "boundaryEvent":
        attrs["attachedToRef"] = ids.get(node.attached_to, node.attached_to)
        if not node.cancel_activity:
            attrs["cancelActivity"] = "false"

    tag = node.type if node.type in NODE_TYPES else "task"
    el = ET.SubElement(parent, _q("bpmn", tag), attrs)

    if node.documentation:
        doc = ET.SubElement(el, _q("bpmn", "documentation"))
        doc.text = node.documentation

    if target == "camunda8":
        ext = None
        if node.type in JOB_TASK_TYPES:
            ext = ET.SubElement(el, _q("bpmn", "extensionElements"))
            ET.SubElement(ext, _q("zeebe", "taskDefinition"),
                          {"type": _sanitize_id(node.id)})
        elif node.type == "userTask":
            ext = ET.SubElement(el, _q("bpmn", "extensionElements"))
            ET.SubElement(ext, _q("zeebe", "userTask"))
    for fid in node.incoming:
        ET.SubElement(el, _q("bpmn", "incoming")).text = fid
    for fid in node.outgoing:
        ET.SubElement(el, _q("bpmn", "outgoing")).text = fid

    doc_ref = doc_ref or {}
    if node.data_in or node.data_out:
        prop_id = f"{ids[node.id]}_prop"
        ET.SubElement(el, _q("bpmn", "property"),
                      {"id": prop_id, "name": "__targetRef_placeholder"})
        for i, ref in enumerate(node.data_in, start=1):
            if ref not in doc_ref:
                continue
            assoc = ET.SubElement(el, _q("bpmn", "dataInputAssociation"),
                                  {"id": f"{ids[node.id]}_din_{i}"})
            ET.SubElement(assoc, _q("bpmn", "sourceRef")).text = doc_ref[ref]
            ET.SubElement(assoc, _q("bpmn", "targetRef")).text = prop_id
        for i, ref in enumerate(node.data_out, start=1):
            if ref not in doc_ref:
                continue
            assoc = ET.SubElement(el, _q("bpmn", "dataOutputAssociation"),
                                  {"id": f"{ids[node.id]}_dout_{i}"})
            ET.SubElement(assoc, _q("bpmn", "targetRef")).text = doc_ref[ref]

    if node.marker in MARKERS and node.type in ACTIVITY_TYPES:
        if node.marker == "loop":
            ET.SubElement(el, _q("bpmn", "standardLoopCharacteristics"))
        else:
            ET.SubElement(el, _q("bpmn", "multiInstanceLoopCharacteristics"),
                          {"isSequential": "true" if node.marker == "mi_sequential" else "false"})

    if node.event_def and node.event_def in EVENT_DEFS and node.type in EVENT_TYPES:
        ET.SubElement(el, _q("bpmn", EVENT_DEFS[node.event_def]),
                      {"id": f"{ids[node.id]}_def"})
    return el


def _emit_flow(parent: ET.Element, flow: Flow, ids: dict[str, str]) -> None:
    attrs = {
        "id": flow.id,
        "sourceRef": ids.get(flow.source, flow.source),
        "targetRef": ids.get(flow.target, flow.target),
    }
    if flow.name:
        attrs["name"] = flow.name
    el = ET.SubElement(parent, _q("bpmn", "sequenceFlow"), attrs)
    if flow.condition:
        cond = ET.SubElement(el, _q("bpmn", "conditionExpression"),
                             {_q("xsi", "type"): "bpmn:tFormalExpression"})
        cond.text = flow.condition
    if flow.is_default:
        source_el = parent.find(f".//*[@id='{attrs['sourceRef']}']")
        if source_el is not None:
            source_el.set("default", flow.id)


# --------------------------------------------------------------------------- #
# Точка входа
# --------------------------------------------------------------------------- #

def convert(data: bytes, *, sheet: str | None = None, strict: bool = False,
            executable: bool = False, target: str = "plain") -> Result:
    """Единственная публичная функция. Ничего не читает и не пишет на диск."""
    nodes, flows, msgs, issues = read_table(data, sheet=sheet)
    by_id, more = validate(nodes, flows, msgs)
    issues += more

    valid_flows = [
        f for f in flows
        if f.source in by_id and f.target in by_id
        and by_id[f.source].pool == by_id[f.target].pool
        and by_id[f.source].type != "endEvent"
        and by_id[f.target].type not in {"startEvent", "boundaryEvent"}
    ]
    valid_msgs = [
        m for m in msgs
        if m.source in by_id and m.target in by_id
        and by_id[m.source].pool != by_id[m.target].pool
    ]

    if target == "camunda8":
        executable = True
        issues += _camunda_rules(by_id, valid_flows)

    has_errors = any(i.level == "error" for i in issues)
    blocked = has_errors or (strict and issues)
    stats = {
        "nodes": len(by_id),
        "flows": len(valid_flows),
        "message_flows": len(valid_msgs),
        "pools": len({n.pool for n in by_id.values()}),
        "errors": sum(1 for i in issues if i.level == "error"),
        "warnings": sum(1 for i in issues if i.level == "warning"),
    }
    if blocked:
        return Result(ok=False, xml=None, issues=issues, stats=stats)

    xml = build_xml(by_id, valid_flows, valid_msgs,
                    executable=executable, target=target)
    return Result(ok=True, xml=xml, issues=issues, stats=stats)


def _camunda_rules(by_id: dict[str, Node], flows: list[Flow]) -> list[Issue]:
    """Проверки и правки, специфичные для Camunda 8 / Zeebe."""
    issues: list[Issue] = []
    starts: dict[str, list[Node]] = {}

    for node in by_id.values():
        if node.marker == "loop":
            issues.append(Issue("error", node.row,
                                f"marker 'loop' у {node.id!r}: Zeebe не поддерживает "
                                f"standardLoopCharacteristics, используйте mi_parallel/mi_sequential"))
        if node.type == "inclusiveGateway" and len(node.incoming) > 1:
            issues.append(Issue("error", node.row,
                                f"сходящийся inclusiveGateway {node.id!r} не поддерживается "
                                f"Camunda 8; замените на parallel/exclusive"))
        if node.type == "boundaryEvent" and not node.event_def:
            issues.append(Issue("error", node.row,
                                f"у boundaryEvent {node.id!r} не задан event_def"))
        if node.type in JOB_TASK_TYPES:
            issues.append(Issue("warning", node.row,
                                f"для {node.id!r} подставлен zeebe:taskDefinition type="
                                f"{_sanitize_id(node.id)!r} — проверьте перед развёртыванием"))
        if node.event_def == "message" or node.type == "receiveTask":
            issues.append(Issue("warning", node.row,
                                f"{node.id!r} ждёт сообщение: bpmn:message и zeebe:subscription "
                                f"нужно дописать в Modeler вручную"))
        if node.type == "startEvent" and not node.event_def and not node.parent_id:
            starts.setdefault(node.pool, []).append(node)

    for pool, nodes in starts.items():
        if len(nodes) > 1:
            issues.append(Issue("error", nodes[1].row,
                                f"в пуле {pool!r} несколько none-стартов "
                                f"({', '.join(n.id for n in nodes)}); Camunda допускает один"))

    for flow in flows:
        if flow.condition and not flow.condition.lstrip().startswith("="):
            flow.condition = "= " + flow.condition.strip()
            issues.append(Issue("warning", flow.row,
                                f"условие потока {flow.id!r} дополнено префиксом '=' "
                                f"по правилам FEEL"))
    return issues
