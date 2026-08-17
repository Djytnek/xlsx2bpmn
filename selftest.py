# -*- coding: utf-8 -*-
"""Проверка на встроенном шаблоне. Запуск: python selftest.py"""
import io
import xml.etree.ElementTree as ET
from pathlib import Path

from xlsx2bpmn import apply_layout, convert, layout_process, to_table, to_workbook
from xlsx2bpmn.layout import DI, B

TEMPLATE = Path(__file__).parent / "xlsx2bpmn" / "template.xlsx"


def main() -> int:
    result = convert(TEMPLATE.read_bytes())
    assert result.ok, [str(i) for i in result.issues]
    print(f"конвертация: {result.stats}")

    xml, warnings = apply_layout(result.xml, layout_process)
    root = ET.fromstring(xml)
    plane = root.find(f".//{DI}BPMNPlane")
    shapes = [e for e in plane if e.tag == f"{DI}BPMNShape"]
    edges = [e for e in plane if e.tag == f"{DI}BPMNEdge"]
    print(f"раскладка: фигур {len(shapes)}, стрелок {len(edges)}")

    semantic = {n.get("id") for p in root.findall(f"{B}process") for n in p
                if n.tag.split("}")[1] not in ("laneSet", "dataObject")}
    drawn = {e.get("bpmnElement") for e in shapes + edges}
    missing = semantic - drawn
    assert not missing, f"без координат: {missing}"
    print("все элементы получили координаты")

    objects = {n.get("id") for p in root.findall(f"{B}process")
               for n in p.findall(f"{B}dataObject")}
    bad = [e for e in shapes if e.get("bpmnElement") in objects]
    assert not bad, "у dataObject не должно быть фигуры"
    print("документы отрисованы корректно")

    again, _ = apply_layout(result.xml, layout_process)
    assert again == xml, "результат недетерминирован"
    print("результат детерминирован")

    # --- обратный ход: круг должен замыкаться ------------------------------
    back = to_table(xml)
    assert back.ok, [str(i) for i in back.issues]
    assert not back.errors, [str(i) for i in back.errors]
    print(f"разбор обратно: строк {back.stats['nodes']}, "
          f"колонок {back.stats['columns']}, потерь {len(back.warnings)}")

    rebuilt = convert(back.table.encode("utf-8"))
    assert rebuilt.ok, [str(i) for i in rebuilt.issues]

    def semantic(text: str) -> str:
        node = ET.fromstring(text)
        for diagram in node.findall(f"{DI}BPMNDiagram"):
            node.remove(diagram)
        ET.indent(node, space="  ")
        return ET.tostring(node, encoding="unicode")

    assert semantic(result.xml) == semantic(rebuilt.xml), "круг исказил схему"
    print("таблица -> схема -> таблица -> схема: смысл сохранён")

    twice, _ = apply_layout(rebuilt.xml, layout_process)
    assert to_table(twice).table == back.table, "таблица не является неподвижной точкой"
    print("таблица — неподвижная точка")

    book = to_workbook(back, TEMPLATE.read_bytes())
    from_book = convert(book)
    assert from_book.ok, [str(i) for i in from_book.issues]
    assert from_book.stats["nodes"] == result.stats["nodes"], "в xlsx потерялись строки"
    print("выгрузка в xlsx читается обратно")

    check_subprocesses()

    for w in warnings:
        print(f"  предупреждение: {w}")
    print("\nВСЁ ХОРОШО")
    return 0


SUBPROCESS_TABLE = """id|name|type|pool|lane|parent_id|next
s|Заявка|startEvent|ООО|Снабжение||sub
sub|Проверка поставщика|subProcess|ООО|Снабжение||e
c_s|Начать|startEvent|ООО||sub|c_t
c_t|Запросить документы|userTask|ООО||sub|deep
deep|Экспертиза|subProcess|ООО||sub|c_e
d_s|Старт|startEvent|ООО||deep|d_t
d_t|Проверить устав|userTask|ООО||deep|d_e
d_e|Готово|endEvent|ООО||deep|
c_e|Проверка завершена|endEvent|ООО||sub|
e|Закупка проведена|endEvent|ООО|Снабжение||"""


def check_subprocesses() -> None:
    """Вложенное содержимое должно и разбираться, и получать координаты."""
    result = convert(SUBPROCESS_TABLE.encode("utf-8"))
    assert result.ok, [str(i) for i in result.issues]
    xml, _ = apply_layout(result.xml, layout_process)
    root = ET.fromstring(xml)

    inner = {"startEvent", "endEvent", "userTask", "subProcess"}
    semantic = {n.get("id") for p in root.findall(f"{B}process") for n in p.iter()
                if n.tag.split("}")[1] in inner}
    drawn: set[str] = set()
    planes = root.findall(f".//{DI}BPMNPlane")
    for plane in planes:
        drawn |= {e.get("bpmnElement") for e in plane}

    missing = semantic - drawn
    assert not missing, f"внутри субпроцессов без координат: {missing}"
    assert len(planes) == 3, f"ожидалось 3 полотна, получено {len(planes)}"
    print(f"субпроцессы: {len(semantic)} элементов, полотен {len(planes)}, "
          f"все с координатами")

    back = to_table(xml)
    assert convert(back.table.encode("utf-8")).ok, "таблица с субпроцессами не собралась"
    assert {r["id"] for r in back.rows} == semantic, "при разборе потерялись элементы"
    nested = {r["id"] for r in back.rows if r["parent_id"]}
    assert nested == {"c_s", "c_t", "deep", "d_s", "d_t", "d_e", "c_e"}, \
        f"неверные parent_id: {nested}"
    print("субпроцессы: разбор обратно сохраняет вложенность")

    # один уровень достаётся отдельно
    only = to_table(xml, level="deep")
    assert {r["id"] for r in only.rows} == {"d_s", "d_t", "d_e"}, \
        f"уровень 'deep' достался неверно: {[r['id'] for r in only.rows]}"
    print("субпроцессы: уровень достаётся отдельной таблицей")

    # книга раскладывается по листам и читается обратно без правок
    book = to_workbook(to_table(xml, full=True), TEMPLATE.read_bytes())
    from openpyxl import load_workbook

    names = load_workbook(io.BytesIO(book)).sheetnames
    assert {"sub", "deep"} <= set(names), f"нет листов субпроцессов: {names}"
    again = convert(book)
    assert again.ok, [str(i) for i in again.issues]
    twice, _ = apply_layout(again.xml, layout_process)
    assert to_table(twice).table == back.table, "книга с листами исказила схему"
    print(f"субпроцессы: книга разложена по листам {names} и читается обратно")


if __name__ == "__main__":
    raise SystemExit(main())
